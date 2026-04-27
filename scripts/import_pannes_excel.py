#!/usr/bin/env python3
"""
import_pannes_excel.py — Script d'import des fiches de panne depuis Excel

Usage :
    python scripts/import_pannes_excel.py <chemin_vers_fichier.xlsx>

Feuille attendue : "DONNÉES"
Colonnes attendues (insensibles à la casse, espaces tolérés) :
  - Date panne (ou "Date") : date de la panne
  - Immatriculation (ou "Immat") : immatriculation du véhicule
  - Société (ou "Soc") : TRAFRIC SARL / BAJ SERVICES SA
  - Service demandé (ou "Service") : description libre
  - Pièces commandées (ou "Pièces") : format libre
  - Statut : en_attente / en_cours / cloture (optionnel, défaut : en_attente)
  - N° moteur (optionnel)
  - Km départ (optionnel)
  - Km fin (optionnel)
  - Mécaniciens (format "Henry / Nangmo") : acteurs rôle mécanicien
  - Chauffeur (format "Prénom Nom") : acteur rôle chauffeur
  - Responsable Opérations (format "Prénom Nom") : acteur rôle responsable

Dépendance externe : openpyxl
    pip install openpyxl
"""

import sys
import os
import re
from datetime import date, datetime
from typing import Optional

# Ajout du répertoire racine du projet au sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

try:
    import openpyxl
except ImportError:
    print("❌ La dépendance 'openpyxl' est absente.")
    print("   Installez-la avec : pip install openpyxl")
    sys.exit(1)

from sqlalchemy.orm import Session
from app.db.database import SessionLocal
from app.models.model_pannes import FichePanne, Acteur, FicheActeur
from sqlalchemy import text


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalize_col(name: str) -> str:
    """Normalise un en-tête de colonne pour la recherche (minuscules, sans espaces)."""
    if not name:
        return ''
    return re.sub(r'\s+', '', str(name).lower().strip())


def _find_col(headers: dict, *candidates: str) -> Optional[int]:
    """Trouve l'index de colonne parmi plusieurs noms possibles."""
    for candidate in candidates:
        normalized = _normalize_col(candidate)
        if normalized in headers:
            return headers[normalized]
    return None


def _parse_date(value) -> Optional[date]:
    """Parse une date depuis diverses représentations Excel."""
    if value is None:
        return None
    if isinstance(value, (datetime,)):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def _parse_acteurs(raw: str) -> list[str]:
    """
    Parse une cellule 'Henry / Nangmo' ou 'Henry/Nangmo' ou 'Henry'
    et retourne une liste de noms normalisés.
    """
    if not raw:
        return []
    # Séparateurs acceptés : / ou , ou ;
    parts = re.split(r'[/,;]', str(raw))
    return [p.strip() for p in parts if p.strip()]


def _get_or_create_acteur(db: Session, nom_complet: str, role: str,
                           acteurs_cache: dict) -> Optional[Acteur]:
    """
    Cherche un acteur par son nom_complet + rôle dans le cache, puis en base.
    Crée l'acteur s'il n'existe pas.
    """
    key = (nom_complet.lower(), role)
    if key in acteurs_cache:
        return acteurs_cache[key]

    # Décomposition prénom / nom
    parts = nom_complet.split()
    if len(parts) >= 2:
        prenom = parts[0].capitalize()
        nom = ' '.join(parts[1:]).upper()
    else:
        prenom = None
        nom = nom_complet.upper()

    # Recherche en base (insensible à la casse)
    from sqlalchemy import func, or_
    acteur = (
        db.query(Acteur)
        .filter(func.lower(Acteur.nom) == nom.lower())
        .filter(Acteur.role == role)
        .first()
    )

    if not acteur:
        acteur = Acteur(nom=nom, prenom=prenom, role=role, actif=True)
        db.add(acteur)
        db.flush()
        acteurs_cache[key] = acteur
        acteurs_cache['__new__'] = acteurs_cache.get('__new__', 0) + 1

    acteurs_cache[key] = acteur
    return acteur


def _get_next_numero(db: Session) -> int:
    result = db.execute(
        text("SELECT COALESCE(MAX(numero_fiche), 0) + 1 FROM fiches_pannes")
    ).scalar()
    return result


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Usage : python scripts/import_pannes_excel.py <fichier.xlsx>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.isfile(xlsx_path):
        print(f"❌ Fichier introuvable : {xlsx_path}")
        sys.exit(1)

    print(f"📂 Lecture du fichier : {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Recherche de la feuille "DONNÉES" (insensible à la casse)
    sheet = None
    for name in wb.sheetnames:
        if name.strip().upper() in ('DONNÉES', 'DONNEES', 'DATA', 'FICHE'):
            sheet = wb[name]
            break

    if sheet is None:
        # Fallback : première feuille
        sheet = wb.active
        print(f"⚠️  Feuille 'DONNÉES' non trouvée — utilisation de la feuille '{sheet.title}'")
    else:
        print(f"✅ Feuille trouvée : '{sheet.title}'")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("❌ Feuille vide.")
        sys.exit(1)

    # Construction du mapping colonne → index depuis la 1ère ligne (en-têtes)
    headers_raw = rows[0]
    headers = {_normalize_col(str(h)): i for i, h in enumerate(headers_raw) if h is not None}
    print(f"   Colonnes détectées : {list(headers.keys())}")

    # Mapping des colonnes
    COL_DATE = _find_col(headers, 'datepanne', 'date', 'datepanne')
    COL_IMMAT = _find_col(headers, 'immatriculation', 'immat', 'véhicule', 'vehicule')
    COL_SOCIETE = _find_col(headers, 'société', 'societe', 'soc', 'entreprise')
    COL_SERVICE = _find_col(headers, 'servicedemandé', 'servicedemande', 'service', 'panne', 'description')
    COL_PIECES = _find_col(headers, 'piècescommandées', 'piecescommandees', 'pièces', 'pieces', 'piècescom')
    COL_STATUT = _find_col(headers, 'statut', 'etat', 'état')
    COL_MOTEUR = _find_col(headers, 'n°moteur', 'nmoteur', 'nomoteur', 'moteur')
    COL_KM_DEP = _find_col(headers, 'kmdépart', 'kmdepart', 'km_départ', 'km_depart', 'kmdep')
    COL_KM_FIN = _find_col(headers, 'kmfin', 'km_fin')
    COL_MECAN = _find_col(headers, 'mécaniciens', 'mecaniciens', 'mécanicien', 'mecanicien', 'mec')
    COL_CHAUF = _find_col(headers, 'chauffeur', 'driver', 'conducteur')
    COL_RESP = _find_col(headers, 'responsableopérations', 'responsableoperations',
                         'responsable', 'resp', 'chefopérations')

    if COL_IMMAT is None:
        print("❌ Colonne 'Immatriculation' introuvable. Arrêt.")
        sys.exit(1)

    # Compteurs
    fiches_crees = 0
    fiches_ignorees = 0
    acteurs_crees = 0
    erreurs = []

    acteurs_cache: dict = {}

    db: Session = SessionLocal()
    try:
        data_rows = rows[1:]  # Ignorer la ligne d'en-têtes
        total = len([r for r in data_rows if any(c is not None for c in r)])
        print(f"\n🔄 Import de {total} ligne(s)...\n")

        for row_idx, row in enumerate(data_rows, start=2):
            # Ignorer les lignes vides
            if all(c is None for c in row):
                continue

            immat = str(row[COL_IMMAT]).strip() if COL_IMMAT is not None and row[COL_IMMAT] else None
            if not immat or immat.lower() in ('none', 'n/a', '-', ''):
                fiches_ignorees += 1
                continue

            # Date panne
            date_panne = None
            if COL_DATE is not None:
                date_panne = _parse_date(row[COL_DATE])
            if date_panne is None:
                date_panne = date.today()  # Fallback

            # Société
            societe = str(row[COL_SOCIETE]).strip() if COL_SOCIETE is not None and row[COL_SOCIETE] else 'TRAFRIC SARL'

            # Statut
            statut_raw = str(row[COL_STATUT]).strip().lower() if COL_STATUT is not None and row[COL_STATUT] else 'en_attente'
            statut_map = {
                'en attente': 'en_attente', 'en_attente': 'en_attente', 'attente': 'en_attente',
                'en cours': 'en_cours', 'en_cours': 'en_cours', 'cours': 'en_cours',
                'clôturé': 'cloture', 'cloture': 'cloture', 'cloturé': 'cloture',
                'clôturée': 'cloture', 'terminé': 'cloture', 'termine': 'cloture',
            }
            statut = statut_map.get(statut_raw, 'en_attente')

            try:
                numero = _get_next_numero(db)

                fiche = FichePanne(
                    numero_fiche=numero,
                    date_panne=date_panne,
                    immatriculation=immat,
                    numero_moteur=(str(row[COL_MOTEUR]).strip()
                                   if COL_MOTEUR is not None and row[COL_MOTEUR] else None),
                    societe=societe,
                    km_depart=(int(row[COL_KM_DEP])
                                if COL_KM_DEP is not None and row[COL_KM_DEP] else None),
                    km_fin=(int(row[COL_KM_FIN])
                             if COL_KM_FIN is not None and row[COL_KM_FIN] else None),
                    service_demande=(str(row[COL_SERVICE]).strip()
                                     if COL_SERVICE is not None and row[COL_SERVICE] else None),
                    pieces_commandees=(str(row[COL_PIECES]).strip()
                                       if COL_PIECES is not None and row[COL_PIECES] else None),
                    statut=statut,
                    created_by_name='import_excel',
                )
                db.add(fiche)
                db.flush()

                # Mécaniciens
                if COL_MECAN is not None and row[COL_MECAN]:
                    noms = _parse_acteurs(str(row[COL_MECAN]))
                    for nom in noms:
                        acteur = _get_or_create_acteur(db, nom, 'mecanicien', acteurs_cache)
                        if acteur:
                            liaison = FicheActeur(
                                fiche_id=fiche.id,
                                acteur_id=acteur.id,
                                role_sur_fiche='mecanicien',
                            )
                            db.add(liaison)

                # Chauffeur
                if COL_CHAUF is not None and row[COL_CHAUF]:
                    noms = _parse_acteurs(str(row[COL_CHAUF]))
                    for nom in noms:
                        acteur = _get_or_create_acteur(db, nom, 'chauffeur', acteurs_cache)
                        if acteur:
                            liaison = FicheActeur(
                                fiche_id=fiche.id,
                                acteur_id=acteur.id,
                                role_sur_fiche='chauffeur',
                            )
                            db.add(liaison)

                # Responsable
                if COL_RESP is not None and row[COL_RESP]:
                    noms = _parse_acteurs(str(row[COL_RESP]))
                    for nom in noms:
                        acteur = _get_or_create_acteur(db, nom, 'responsable', acteurs_cache)
                        if acteur:
                            liaison = FicheActeur(
                                fiche_id=fiche.id,
                                acteur_id=acteur.id,
                                role_sur_fiche='responsable',
                            )
                            db.add(liaison)

                db.commit()
                fiches_crees += 1

                if fiches_crees % 50 == 0:
                    print(f"   … {fiches_crees} fiches importées")

            except Exception as e:
                db.rollback()
                erreurs.append(f"Ligne {row_idx} ({immat}): {e}")
                fiches_ignorees += 1

        acteurs_crees = acteurs_cache.get('__new__', 0)

    finally:
        db.close()

    # Résumé
    print("\n" + "=" * 50)
    print("📊 RÉSUMÉ DE L'IMPORT")
    print("=" * 50)
    print(f"  ✅ Fiches importées  : {fiches_crees}")
    print(f"  👤 Acteurs créés     : {acteurs_crees}")
    print(f"  ⏭️  Lignes ignorées   : {fiches_ignorees}")
    if erreurs:
        print(f"\n  ❌ Erreurs ({len(erreurs)}) :")
        for e in erreurs[:20]:
            print(f"     {e}")
        if len(erreurs) > 20:
            print(f"     ... et {len(erreurs) - 20} autres erreurs")
    print("=" * 50)


if __name__ == '__main__':
    main()
